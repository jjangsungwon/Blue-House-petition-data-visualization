from openpyxl import load_workbook
from konlpy.tag import Kkma
import collections
import numpy as np
import copy
import matplotlib.pyplot as plt
from wordcloud import WordCloud
from os import path
from PIL import Image
import numpy as np
import os

read_workbook = load_workbook("./bluehouse.xlsx")
read_cell = read_workbook.active

# 엑셀 150줄까지 읽는다.
list_excel = []
for i in range(1, 151):
    list_excel.append(read_cell.cell(i, 1).value)

# 자연어 처리
kkma = Kkma()
list_temp = []

for row in list_excel:
    list_temp = list_temp + kkma.nouns(row)

# 한글자로 이루어진 것은 제거한다(정확하지 않은 단어일 확률이 높다)
result_list = []
for i in list_temp:
    if len(i) == 1:
        continue
    else:
        result_list.append(i)

copy_list = copy.deepcopy(result_list)
# print(collections.Counter(result_list))
# print(collections.Counter(result_list).most_common(5))  # 상위 5개
result_list = dict(collections.Counter(result_list).most_common(20))  # 상위 20개

# 시각화(막대 그래프)
# 문자, 숫자 분리 (그래프를 그리기 위해서)
list_string = []
list_number = []
for s, n in result_list.items():
    list_string.append(s)
    list_number.append(n)

# 한글 깨짐 방지
plt.rcParams["font.family"] = 'Malgun Gothic'

label = list_string
index = np.arange(len(label))
plt.bar(index, list_number)
plt.title("my title")
plt.xlabel("x-label")
plt.ylabel("y-label")
plt.xticks(index, label)
plt.xticks(rotation=60)
plt.show()

# 시각화(wordcloud) - 문자열 형태로 넘겨주면 알아서 잘라서 만들어준다.
last_text = ""
for i in copy_list:
    last_text = last_text + " " + i
d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()
mask = np.array(Image.open(path.join(d, './img/cloud.png')))  # 마스크 없으면 기본 wordcloud 제공
#wordcloud = WordCloud(font_path="C:/WINDOWS/FONTS/NANUMGOTHIC.TTF", background_color="white", mask=mask, ).generate(last_text)
wordcloud = WordCloud(font_path="C:/WINDOWS/FONTS/NANUMGOTHIC.TTF").generate(last_text)
plt.figure(figsize=(12, 12))
plt.imshow(wordcloud, interpolation="bilinear")
plt.axis("off")
plt.show()
